import customtkinter as ctk
import tkinter as tk
import random
import json
import csv
import os
import math
import datetime

# --- Инициализация окна ---
app = ctk.CTk()
app.geometry("500x750")
app.minsize(420, 600)
app.title("100 Заключённых")
app.configure(fg_color="#000000")
app.resizable(True, True)
# Sounds removed.


# --- Цвета ---
G        = "#1eb852"
G_BRIGHT = "#2eff6e"
G_DIM    = "#0d3d22"
BG       = "#000000"
color_green = "#1eb852"
color_red   = "#e62e2e"

# --- Шрифты ---
font_title_number = ctk.CTkFont(family="Arial", size=42, weight="bold")
font_title_text   = ctk.CTkFont(family="Arial", size=36, weight="bold")
font_button       = ctk.CTkFont(family="Arial", size=18, weight="bold")

# =============================================
# СИСТЕМА ЯЗЫКОВ
# =============================================

TRANSLATIONS = {
    "ru": {
        "title_num": "100",
        "title_word": "Заключенных",
        "menu_play": "Играть",
        "menu_stats": "Статистика",
        "menu_history": "История игр",
        "menu_settings": "Настройки",
        "menu_export": "Экспорт",
        "menu_exit": "Выйти",
        "settings_title": "Настройки",
        "settings_lang": "Настройки Язык",
        "settings_reset": "Сброс Данных",
        "settings_rewrite": "Перезапись",
        "settings_rules": "Правила игры",
        "lang_english": "English",
        "lang_russian": "Русский",
        "lang_kyrgyz": "Кыргызча",
        "reset_warning": "Внимание хотим вас оповестить о том что после сброса удаленные данные буду утеряны,\nтак что советуем экспортировать ваши данные",
        "reset_export_btn": "☁  Экспорт в Данных",
        "reset_btn": "🗑  Сброс Данных",
        "reset_done": "✓ Данные успешно сброшены",
        "rules_text": (
            "Правила игры — 100 Заключенных\n\n"
            "Условие:\n"
            "• 100 заключённых пронумерованы от 1 до 100.\n"
            "• В комнате стоят 100 пронумерованных коробок.\n"
            "• В каждой коробке случайно лежит один номер (от 1 до 100).\n\n"
            "Цель:\n"
            "• Каждый заключённый должен найти коробку со своим номером.\n"
            "• Каждому разрешено открыть не более 50 коробок.\n"
            "• Все 100 заключённых должны найти свой номер — иначе все погибают.\n\n"
            "Стратегия циклов:\n"
            "• Заключённый начинает с коробки под своим номером.\n"
            "• Открывает её, смотрит на число внутри.\n"
            "• Переходит к коробке с этим числом.\n"
            "• Повторяет, пока не найдёт свой номер или не исчерпает 50 попыток.\n\n"
            "Вероятность выживания:\n"
            "• Случайная стратегия: ≈ (1/2)^100 ≈ 0%\n"
            "• Стратегия циклов: ≈ 31.18%\n\n"
            "Математическое обоснование:\n"
            "• Стратегия провалится только если в перестановке есть цикл длиннее 50.\n"
            "• Вероятность этого: ln(2) ≈ 0.6931, то есть успех ≈ 1 - ln(2) ≈ 31%."
        ),
        "export_title": "Экспорт Данных",
        "export_sub": "Экспортируйте историю игр и статистику в форматах CSV, JSON и Excel",
        "export_stats_title": "Статистика",
        "export_stats_desc": "Экспорт общей статистики, включая данные по стратегиям, процент успеха и другие показатели",
        "export_history_title": "История игр",
        "export_history_desc": "Экспорт списка всех игр с основной информацией: дата, стратегия, результат, количество выживших",
        "export_detail_title": "Детальная история",
        "export_detail_desc": "Экспорт полной истории с детальной информацией о каждом заключённом: попытки, открытые коробки, результаты. Каждая игра экспортируется на отдельный лист.",
        "btn_csv": "Экспорт в CSV",
        "btn_excel": "Экспорт в Excel",
        "btn_json": "Экспорт в JSON",
        "stats_title": "История и Статистика",
        "stats_total": "Всего игр",
        "stats_success": "Успешных игр",
        "stats_rate": "Процент успеха",
        "stats_tab_stats": "Статистика",
        "stats_tab_history": "История игр",
        "stats_compare": "Сравнение стратегий",
        "stats_overall": "Общее соотношение",
        "stats_cyclic": "⟲ Циклическая стратегия",
        "stats_random": "Случайная стратегия",
        "stats_total_g": "Всего игр:",
        "stats_succ": "Успешных:",
        "stats_fail": "Провалов:",
        "stats_success_label": "Успех:",
        "history_win": "Выигрыш",
        "history_lose": "Провал",
        "history_date": "Дата:",
        "history_time": "Время:",
        "btn_again": "Заново",
        "btn_exit_menu": "Назад в меню",
        "btn_exit": "← ВЫЙТИ",
        "game_subtitle": "Классическая математическая\nголоволомка",
        "game_desc": "Каждый заключенный должен найти свой номер\nв коробках. Используя стратегию циклов,\nшанс выживания увеличивается с 0% до 31%",
        "btn_manual": "▶ Начать в Ручную",
        "btn_auto": "▶ Начать в Авто",
        "prisoner_label": "Заключенный #",
        "tries_left": "Попыток\nосталось:",
        "succ_prisoners": "Успешных\nзаключенных",
        "used_tries": "Использовано\nПопыток",
        "game_won": "ВЫ ВЫИГРАЛИ!",
        "game_lost": "ИГРА ПРОИГРАНА",
        "auto_title": "МАТЕМАТИЧЕСКАЯ ГОЛОВОЛОМКА",
        "auto_sub": "Классическая задача о вероятности и выживании • Стратегия циклов • 31% успеха",
        "rules_title_auto": "Правила игры",
        "rules_auto": (
            "• 100 заключенных и 100 коробок с номерами от 1 до 100\n"
            "• В каждой коробке случайно размещен один номер\n"
            "• Каждый заключенный может открыть до 50 коробок\n"
            "• Цель: найти коробку со своим номером\n"
            "• Все выживают, только если КАЖДЫЙ найдет свой номер\n"
            "• Заключенные не могут общаться после начала игры"
        ),
        "sim_settings": "⚙ Настройки симуляции",
        "num_prisoners": "Количество заключенных",
        "num_games": "Количество Игр в быстрой симуляции",
        "choose_strategy": "Выберите стратегию",
        "strat_cyclic": "Стратегия Циклов",
        "strat_random": "Случайная",
        "start_btn": "ПУСК",
        "total_games_res": "Всего игр",
        "success_pct": "Успех %",
        "success_res": "Успешных",
        "fail_res": "Провалов",
        "strategy_cyclic_hist": "Циклическая стратегия",
        "strategy_random_hist": "Случайная стратегия",
        "strategy_manual_hist": "Ручная игра",
        "back_to_menu": "Назад в меню",
        "rewrite_desc": "Эта функция позволяет перезапустить приложение,\nсохранив все текущие данные.",
    },
    "en": {
        "title_num": "100",
        "title_word": "Prisoners",
        "menu_play": "Play",
        "menu_stats": "Statistics",
        "menu_history": "Game History",
        "menu_settings": "Settings",
        "menu_export": "Export",
        "menu_exit": "Exit",
        "settings_title": "Settings",
        "settings_lang": "Language Settings",
        "settings_reset": "Reset Data",
        "settings_rewrite": "Rewrite",
        "settings_rules": "Game Rules",
        "lang_english": "English",
        "lang_russian": "Русский",
        "lang_kyrgyz": "Кыргызча",
        "reset_warning": "Warning: after reset all deleted data will be lost,\nwe recommend exporting your data first.",
        "reset_export_btn": "☁  Export Data",
        "reset_btn": "🗑  Reset Data",
        "reset_done": "✓ Data successfully reset",
        "rules_text": (
            "Game Rules — 100 Prisoners\n\n"
            "Setup:\n"
            "• 100 prisoners are numbered 1 to 100.\n"
            "• A room contains 100 numbered boxes.\n"
            "• Each box contains one random number (1–100).\n\n"
            "Goal:\n"
            "• Each prisoner must find the box with their number.\n"
            "• Each prisoner may open at most 50 boxes.\n"
            "• All 100 must succeed — otherwise everyone dies.\n\n"
            "Loop Strategy:\n"
            "• Start at the box matching your own number.\n"
            "• Open it and note the number inside.\n"
            "• Go to the box with that number.\n"
            "• Repeat until you find your number or run out of tries.\n\n"
            "Survival Probability:\n"
            "• Random strategy: ≈ (1/2)^100 ≈ 0%\n"
            "• Loop strategy: ≈ 31.18%\n\n"
            "Math:\n"
            "• The strategy fails only if a permutation cycle exceeds length 50.\n"
            "• Probability of that: ln(2) ≈ 0.6931, so success ≈ 1 − ln(2) ≈ 31%."
        ),
        "export_title": "Export Data",
        "export_sub": "Export game history and statistics in CSV, JSON and Excel formats",
        "export_stats_title": "Statistics",
        "export_stats_desc": "Export overall statistics including strategy data, success rate and other metrics",
        "export_history_title": "Game History",
        "export_history_desc": "Export a list of all games with basic info: date, strategy, result, survivors",
        "export_detail_title": "Detailed History",
        "export_detail_desc": "Export full history with detailed info about each prisoner: attempts, opened boxes, results. Each game is exported to a separate sheet.",
        "btn_csv": "Export to CSV",
        "btn_excel": "Export to Excel",
        "btn_json": "Export to JSON",
        "stats_title": "History & Statistics",
        "stats_total": "Total games",
        "stats_success": "Successful games",
        "stats_rate": "Success rate",
        "stats_tab_stats": "Statistics",
        "stats_tab_history": "Game History",
        "stats_compare": "Strategy comparison",
        "stats_overall": "Overall ratio",
        "stats_cyclic": "⟲ Loop strategy",
        "stats_random": "Random strategy",
        "stats_total_g": "Total:",
        "stats_succ": "Success:",
        "stats_fail": "Fails:",
        "stats_success_label": "Success:",
        "history_win": "Win",
        "history_lose": "Loss",
        "history_date": "Date:",
        "history_time": "Time:",
        "btn_again": "Replay",
        "btn_exit_menu": "Back to menu",
        "btn_exit": "← EXIT",
        "game_subtitle": "Classic mathematical\npuzzle",
        "game_desc": "Each prisoner must find their number\nin the boxes. Using the loop strategy,\nsurvival chances rise from 0% to 31%",
        "btn_manual": "▶ Play Manually",
        "btn_auto": "▶ Auto Simulate",
        "prisoner_label": "Prisoner #",
        "tries_left": "Tries\nleft:",
        "succ_prisoners": "Successful\nprisoners",
        "used_tries": "Tries\nUsed",
        "game_won": "YOU WIN!",
        "game_lost": "GAME OVER",
        "auto_title": "MATHEMATICAL PUZZLE",
        "auto_sub": "Classic probability and survival problem • Loop strategy • 31% success",
        "rules_title_auto": "Game Rules",
        "rules_auto": (
            "• 100 prisoners and 100 boxes numbered 1 to 100\n"
            "• Each box contains one random number\n"
            "• Each prisoner may open up to 50 boxes\n"
            "• Goal: find the box with your number\n"
            "• All survive only if EVERY prisoner finds their number\n"
            "• Prisoners cannot communicate after the game starts"
        ),
        "sim_settings": "⚙ Simulation Settings",
        "num_prisoners": "Number of prisoners",
        "num_games": "Number of games in quick simulation",
        "choose_strategy": "Choose strategy",
        "strat_cyclic": "Loop Strategy",
        "strat_random": "Random",
        "start_btn": "START",
        "total_games_res": "Total games",
        "success_pct": "Success %",
        "success_res": "Successful",
        "fail_res": "Failed",
        "strategy_cyclic_hist": "Loop strategy",
        "strategy_random_hist": "Random strategy",
        "strategy_manual_hist": "Manual game",
        "back_to_menu": "Back to menu",
        "rewrite_desc": "This function restarts the application\nwhile keeping all current data.",
    },
    "ky": {
        "title_num": "100",
        "title_word": "Туткун",
        "menu_play": "Ойноо",
        "menu_stats": "Статистика",
        "menu_history": "Оюн тарыхы",
        "menu_settings": "Жөндөөлөр",
        "menu_export": "Экспорт",
        "menu_exit": "Чыгуу",
        "settings_title": "Жөндөөлөр",
        "settings_lang": "Тил жөндөөлөрү",
        "settings_reset": "Маалыматты тазалоо",
        "settings_rewrite": "Кайра жазуу",
        "settings_rules": "Оюн эрежелери",
        "lang_english": "English",
        "lang_russian": "Русский",
        "lang_kyrgyz": "Кыргызча",
        "reset_warning": "Эскертүү: тазалагандан кийин жок кылынган маалыматтар жоголот,\nмаалыматтарыңызды алдын ала экспорттоону сунуштайбыз.",
        "reset_export_btn": "☁  Маалыматты экспорттоо",
        "reset_btn": "🗑  Маалыматты тазалоо",
        "reset_done": "✓ Маалыматтар ийгиликтүү тазаланды",
        "rules_text": (
            "Оюн эрежелери — 100 Туткун\n\n"
            "Шарттар:\n"
            "• 100 туткун 1ден 100гө чейин номерленген.\n"
            "• Бөлмөдө 100 номерленген кутуча бар.\n"
            "• Ар бир кутучада бир кокустук номер жатат.\n\n"
            "Максат:\n"
            "• Ар бир туткун өз номери бар кутучаны табышы керек.\n"
            "• Ар бирине 50 кутучага чейин ачууга уруксат берилет.\n"
            "• Баары 100 туткун табышы керек — болбосо бардыгы өлөт.\n\n"
            "Цикл стратегиясы:\n"
            "• Өз номериңдеги кутучадан баштаңыз.\n"
            "• Ачып, ичиндеги санды белгилеңиз.\n"
            "• Ошол санга карата кутучага өтүңүз.\n"
            "• Номериңизди тапмайынча же аракет бүтмөйүнчө кайталаңыз.\n\n"
            "Жашоо мүмкүнчүлүгү:\n"
            "• Туш келди стратегия: ≈ 0%\n"
            "• Цикл стратегиясы: ≈ 31.18%"
        ),
        "export_title": "Маалыматты экспорттоо",
        "export_sub": "Оюн тарыхын жана статистиканы CSV, JSON жана Excel форматтарында экспорттоо",
        "export_stats_title": "Статистика",
        "export_stats_desc": "Жалпы статистиканы экспорттоо: стратегиялар, ийгилик пайызы жана башка көрсөткүчтөр",
        "export_history_title": "Оюн тарыхы",
        "export_history_desc": "Бардык оюндардын тизмесин экспорттоо: дата, стратегия, натыйжа",
        "export_detail_title": "Толук тарых",
        "export_detail_desc": "Ар бир туткун жөнүндө толук маалымат менен тарыхты экспорттоо. Ар бир оюн өзүнчө баракка.",
        "btn_csv": "CSV форматта",
        "btn_excel": "Excel форматта",
        "btn_json": "JSON форматта",
        "stats_title": "Тарых жана Статистика",
        "stats_total": "Бардык оюндар",
        "stats_success": "Ийгиликтүү оюндар",
        "stats_rate": "Ийгилик пайызы",
        "stats_tab_stats": "Статистика",
        "stats_tab_history": "Оюн тарыхы",
        "stats_compare": "Стратегияларды салыштыруу",
        "stats_overall": "Жалпы катыш",
        "stats_cyclic": "⟲ Цикл стратегиясы",
        "stats_random": "Туш келди стратегия",
        "stats_total_g": "Бардыгы:",
        "stats_succ": "Ийгиликтүү:",
        "stats_fail": "Ийгиликсиз:",
        "stats_success_label": "Ийгилик:",
        "history_win": "Жеңиш",
        "history_lose": "Жеңилүү",
        "history_date": "Дата:",
        "history_time": "Убакыт:",
        "btn_again": "Кайра",
        "btn_exit_menu": "Менюга кайтуу",
        "btn_exit": "← ЧЫГУУ",
        "game_subtitle": "Классикалык математикалык\nоюн",
        "game_desc": "Ар бир туткун кутучалардан өз номерин\nтабышы керек. Цикл стратегиясын колдонуу\nменен 0%дан 31%га чейин жогорулайт",
        "btn_manual": "▶ Кол менен баштоо",
        "btn_auto": "▶ Авто баштоо",
        "prisoner_label": "Туткун #",
        "tries_left": "Аракет\nкалды:",
        "succ_prisoners": "Ийгиликтүү\nтуткундар",
        "used_tries": "Колдонулган\nАракеттер",
        "game_won": "ЖЕҢДИҢИЗ!",
        "game_lost": "ОЮН ЖОГОЛДУ",
        "auto_title": "МАТЕМАТИКАЛЫК ОЮН",
        "auto_sub": "Классикалык ыктымалдык маселеси • Цикл стратегиясы • 31% ийгилик",
        "rules_title_auto": "Оюн эрежелери",
        "rules_auto": (
            "• 100 туткун жана 100 кутуча 1ден 100гө чейин\n"
            "• Ар бир кутучада кокустук номер бар\n"
            "• Ар бир туткун 50 кутучага чейин ача алат\n"
            "• Максат: өз номериңди табуу\n"
            "• Баары тирүү калат, эгер ар бири өз номерин тапса\n"
            "• Оюн башталгандан кийин туткундар сүйлөшө алышпайт"
        ),
        "sim_settings": "⚙ Симуляция жөндөөлөрү",
        "num_prisoners": "Туткундардын саны",
        "num_games": "Тез симуляциядагы оюндар саны",
        "choose_strategy": "Стратегия тандоо",
        "strat_cyclic": "Цикл Стратегиясы",
        "strat_random": "Туш келди",
        "start_btn": "БАШТОО",
        "total_games_res": "Бардык оюндар",
        "success_pct": "Ийгилик %",
        "success_res": "Ийгиликтүү",
        "fail_res": "Ийгиликсиз",
        "strategy_cyclic_hist": "Цикл стратегиясы",
        "strategy_random_hist": "Туш келди стратегия",
        "strategy_manual_hist": "Кол менен оюн",
        "back_to_menu": "Менюга кайтуу",
        "rewrite_desc": "Бул функция колдонмону кайра иштетет,\nбардык учурдагы маалыматтарды сактайт.",
    }
}

current_lang = "ru"

def T(key):
    return TRANSLATIONS[current_lang].get(key, TRANSLATIONS["ru"].get(key, key))


# =============================================
# ДАННЫЕ СТАТИСТИКИ
# =============================================

class StatsData:
    history = []
    total_games = 0
    success_games = 0
    success_rate = 0
    cyclic_total = 0
    cyclic_success = 0
    cyclic_fails = 0
    random_total = 0
    random_success = 0
    random_fails = 0

    @classmethod
    def load_db(cls):
        import sqlite3
        try:
            conn = sqlite3.connect('game_data.db')
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS stats
                         (id INTEGER PRIMARY KEY, total_games INT, success_games INT, success_rate INT, 
                          cyclic_total INT, cyclic_success INT, cyclic_fails INT,
                          random_total INT, random_success INT, random_fails INT)''')
            c.execute('''CREATE TABLE IF NOT EXISTS history
                         (id INTEGER PRIMARY KEY AUTOINCREMENT, is_win INT, strategy TEXT, rate INT, date TEXT, time TEXT)''')
            
            c.execute("SELECT * FROM stats WHERE id=1")
            row = c.fetchone()
            if row:
                cls.total_games = row[1]
                cls.success_games = row[2]
                cls.success_rate = row[3]
                cls.cyclic_total = row[4]
                cls.cyclic_success = row[5]
                cls.cyclic_fails = row[6]
                cls.random_total = row[7]
                cls.random_success = row[8]
                cls.random_fails = row[9]
            else:
                c.execute("INSERT INTO stats (id, total_games, success_games, success_rate, cyclic_total, cyclic_success, cyclic_fails, random_total, random_success, random_fails) VALUES (1,0,0,0,0,0,0,0,0,0)")
                conn.commit()
                
            c.execute("SELECT is_win, strategy, rate, date, time FROM history ORDER BY id DESC")
            rows = c.fetchall()
            cls.history = []
            for r in rows:
                cls.history.append({
                    'is_win': bool(r[0]),
                    'strategy': r[1],
                    'rate': r[2],
                    'date': r[3],
                    'time': r[4]
                })
            conn.close()
        except Exception as e:
            print("DB Load Error:", e)

    @classmethod
    def save_db(cls):
        import sqlite3
        try:
            conn = sqlite3.connect('game_data.db')
            c = conn.cursor()
            c.execute('''UPDATE stats SET 
                         total_games=?, success_games=?, success_rate=?,
                         cyclic_total=?, cyclic_success=?, cyclic_fails=?,
                         random_total=?, random_success=?, random_fails=? WHERE id=1''',
                      (cls.total_games, cls.success_games, cls.success_rate,
                       cls.cyclic_total, cls.cyclic_success, cls.cyclic_fails,
                       cls.random_total, cls.random_success, cls.random_fails))
                       
            c.execute("DELETE FROM history")
            for item in reversed(cls.history):
                c.execute("INSERT INTO history (is_win, strategy, rate, date, time) VALUES (?,?,?,?,?)",
                          (int(item['is_win']), item['strategy'], item['rate'], item['date'], item['time']))
            conn.commit()
            conn.close()
        except Exception as e:
            print("DB Save Error:", e)

StatsData.load_db()



def run_simulation(strategy, num_prisoners, num_games):
    success_count = 0
    fail_count = 0
    max_tries = num_prisoners // 2

    for _ in range(num_games):
        boxes = list(range(1, num_prisoners + 1))
        random.shuffle(boxes)
        game_won = True
        if strategy == 'cyclic':
            visited = set()
            max_cycle = 0
            for i in range(1, num_prisoners + 1):
                if i not in visited:
                    current = i
                    cycle_len = 0
                    while current not in visited:
                        visited.add(current)
                        current = boxes[current - 1]
                        cycle_len += 1
                    max_cycle = max(max_cycle, cycle_len)
            if max_cycle > max_tries:
                game_won = False
        else:
            for p in range(1, num_prisoners + 1):
                choices = random.sample(range(num_prisoners), max_tries)
                found = False
                for choice in choices:
                    if boxes[choice] == p:
                        found = True
                        break
                if not found:
                    game_won = False
                    break

        if game_won:
            success_count += 1
        else:
            fail_count += 1

    return success_count, fail_count


# =============================================
# АНИМИРОВАННОЕ ГЛАВНОЕ МЕНЮ (из кода 2)
# =============================================

# Глобальные ссылки на виджеты меню (нужны для смены языка)
menu_canvas   = None
menu_frame    = None
lbl_100       = None
lbl_title     = None
btn_widgets   = []      # список кнопок меню
particles     = []
particle_job  = None    # id after-задачи частиц

# Экраны второго уровня
game_screen         = None
auto_sim_frame      = None
manual_sim_frame    = None
stats_screen_frame  = None
export_screen_frame = None
settings_screen_frame = None


def _build_particles():
    global particles
    particles = []
    for _ in range(70):
        particles.append({
            "x":     random.uniform(0, 2000),
            "y":     random.uniform(0, 1000),
            "speed": random.uniform(0.3, 1.2),
            "size":  random.uniform(1, 3),
            "alpha": random.uniform(0.2, 0.9),
        })


def _animate_particles():
    global particle_job
    if menu_canvas is None:
        return
    try:
        menu_canvas.delete("particle")
    except Exception:
        return
    w = app.winfo_width()  or 500
    h = app.winfo_height() or 750
    for p in particles:
        p["y"] -= p["speed"]
        if p["y"] < -5:
            p["y"] = h + 5
            p["x"] = random.uniform(0, w)
        g_val  = min(255, int(p["alpha"] * 184))
        r_val  = int(p["alpha"] * 30)
        color  = f"#{r_val:02x}{g_val:02x}{r_val:02x}"
        r = p["size"]
        menu_canvas.create_oval(
            p["x"] - r, p["y"] - r,
            p["x"] + r, p["y"] + r,
            fill=color, outline="", tags="particle")
    particle_job = app.after(30, _animate_particles)


def _stop_particles():
    global particle_job
    if particle_job is not None:
        app.after_cancel(particle_job)
        particle_job = None


def _hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def _rgb_to_hex(r, g, b):
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

def _lerp(a, b, t):
    return a + (b - a) * t

def _smoothstep(t):
    return t * t * (3 - 2 * t)


def _make_hover_animation(btn, entering: bool):
    steps = 12
    delay = 12
    if entering:
        start_bg, end_bg     = "#000000", "#0a2918"
        start_text, end_text = G, G_BRIGHT
        start_bord, end_bord = G, G_BRIGHT
    else:
        start_bg, end_bg     = "#0a2918", "#000000"
        start_text, end_text = G_BRIGHT, G
        start_bord, end_bord = G_BRIGHT, G

    s_bg  = _hex_to_rgb(start_bg);   e_bg  = _hex_to_rgb(end_bg)
    s_tx  = _hex_to_rgb(start_text); e_tx  = _hex_to_rgb(end_text)
    s_br  = _hex_to_rgb(start_bord); e_br  = _hex_to_rgb(end_bord)

    def step(i=0):
        if i > steps:
            return
        t    = _smoothstep(i / steps)
        bg   = _rgb_to_hex(*[_lerp(s, e, t) for s, e in zip(s_bg, e_bg)])
        text = _rgb_to_hex(*[_lerp(s, e, t) for s, e in zip(s_tx, e_tx)])
        bord = _rgb_to_hex(*[_lerp(s, e, t) for s, e in zip(s_br, e_br)])
        try:
            btn.configure(fg_color=bg, text_color=text, border_color=bord)
        except Exception:
            pass
        app.after(delay, lambda: step(i + 1))

    step()


def _make_click_animation(btn, command=None):
    def flash(i=0):
        colors = [G_BRIGHT, "#ffffff", G_BRIGHT, G]
        if i < len(colors):
            try:
                btn.configure(text_color=colors[i], border_color=colors[i])
            except Exception:
                pass
            app.after(60, lambda: flash(i + 1))
        else:
            if command:
                command()
    flash()


def transition_to(target_func, *args, **kwargs):
    overlay = ctk.CTkToplevel(app)
    overlay.overrideredirect(True)
    overlay.attributes("-alpha", 0.0)
    overlay.configure(fg_color="#000000")
    overlay.geometry(f"{app.winfo_width()}x{app.winfo_height()}+{app.winfo_rootx()}+{app.winfo_rooty()}")
    overlay.lift()

    def fade_out(step=0, total=12):
        alpha = _smoothstep(step / total)
        try:
            overlay.attributes("-alpha", alpha)
        except Exception:
            return
        if step < total:
            app.after(10, lambda: fade_out(step + 1, total))
        else:
            target_func(*args, **kwargs)
            fade_in()

    def fade_in(step=0, total=12):
        alpha = 1.0 - _smoothstep(step / total)
        try:
            overlay.attributes("-alpha", alpha)
        except Exception:
            return
        if step < total:
            app.after(10, lambda: fade_in(step + 1, total))
        else:
            try:
                overlay.destroy()
            except Exception:
                pass

    fade_out()

MENU_BUTTONS_DEF = [
    ("▷",  "menu_play",     lambda: transition_to(show_game_screen)),
    ("▦",  "menu_stats",    lambda: transition_to(show_stats_screen, "stats")),
    ("↺",  "menu_history",  lambda: transition_to(show_stats_screen, "history")),
    ("⚙",  "menu_settings", lambda: transition_to(show_settings_screen)),
    ("☁",  "menu_export",   lambda: transition_to(show_export_screen)),
    ("→",  "menu_exit",     lambda: app.quit()),
]


def build_menu(first_time=True):
    """Создаёт (или пересоздаёт) главное меню с анимированным фоном."""
    global menu_canvas, menu_frame, lbl_100, lbl_title, btn_widgets

    app.minsize(420, 600)

    # Удаляем старые виджеты если есть
    for w in app.winfo_children():
        if getattr(w, "is_overlay", False): continue
        try:
            w.destroy()
        except Exception:
            pass

    _build_particles()

    # Canvas для частиц
    menu_canvas = tk.Canvas(app, bg=BG, highlightthickness=0)
    menu_canvas.place(x=0, y=0, relwidth=1, relheight=1)

    # Центральный фрейм
    menu_frame = ctk.CTkFrame(app, fg_color="transparent")
    menu_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Заголовок
    lbl_100 = ctk.CTkLabel(
        menu_frame, text="100",
        font=ctk.CTkFont(family="Arial Black", size=72, weight="bold"),
        text_color=G)
    lbl_100.pack()

    lbl_title = ctk.CTkLabel(
        menu_frame, text=T("title_word"),
        font=ctk.CTkFont(family="Arial Black", size=44, weight="bold"),
        text_color=G)
    lbl_title.pack(pady=(0, 40))

    btn_widgets = []

    for icon, tkey, cmd in MENU_BUTTONS_DEF:
        text = f"  {icon}    {T(tkey)}"
        btn = ctk.CTkButton(
            menu_frame,
            text=text,
            font=ctk.CTkFont(family="Arial", size=18, weight="bold"),
            text_color=G,
            fg_color="#000000",
            border_color=G,
            border_width=2,
            corner_radius=14,
            hover=False,
            width=290,
            height=52,
            anchor="w",
            command=lambda c=cmd, b=None: None   # заглушка — перезапишем ниже
        )
        btn.pack(pady=7)

        # Нужна замкнутая ссылка на btn
        _cmd = cmd
        _btn = btn
        btn.configure(command=lambda b=_btn, c=_cmd: _make_click_animation(b, c))
        btn.bind("<Enter>", lambda e, b=_btn: _make_hover_animation(b, True))
        btn.bind("<Leave>", lambda e, b=_btn: _make_hover_animation(b, False))

        btn_widgets.append(btn)

    # Скрываем всё для анимации появления
    if first_time:
        lbl_100.configure(text_color="#000000")
        lbl_title.configure(text_color="#000000")
        for b in btn_widgets:
            b.configure(text_color="#000000", border_color="#000000", fg_color="#000000")

    # Запуск анимаций
    _animate_particles()
    
    if first_time:
        app.after(100, lambda: _fade_in_window())
        app.after(500, lambda: _animate_title_in())
    else:
        lbl_100.configure(text_color=G)
        lbl_title.configure(text_color=G)
        for b in btn_widgets:
            b.configure(text_color=G, border_color=G, fg_color="#000000")


def _fade_in_window(step=0, total=25):
    alpha = _smoothstep(step / total)
    app.attributes("-alpha", alpha)
    if step < total:
        app.after(18, lambda: _fade_in_window(step + 1, total))


def _animate_title_in(step=0, total=20):
    t     = _smoothstep(step / total)
    g_val = int(t * 184)
    r_val = int(t * 30)
    b_val = int(t * 82)
    color = f"#{r_val:02x}{g_val:02x}{b_val:02x}"
    try:
        lbl_100.configure(text_color=color)
        lbl_title.configure(text_color=color)
    except Exception:
        return
    if step < total:
        app.after(20, lambda: _animate_title_in(step + 1, total))
    else:
        _animate_buttons_in(0)


def _animate_buttons_in(idx):
    if idx >= len(btn_widgets):
        return
    btn = btn_widgets[idx]

    def btn_fade(step=0, total=15):
        t     = _smoothstep(step / total)
        g_val = int(t * 184)
        r_val = int(t * 30)
        b_val = int(t * 82)
        color = f"#{r_val:02x}{g_val:02x}{b_val:02x}"
        try:
            btn.configure(text_color=color, border_color=color)
        except Exception:
            pass
        if step < total:
            app.after(16, lambda: btn_fade(step + 1, total))
        else:
            btn.configure(text_color=G, border_color=G, fg_color="#000000")
            app.after(60, lambda: _animate_buttons_in(idx + 1))

    btn_fade()


def back_to_menu(frame_to_hide=None):
    """Возврат в главное меню из любого экрана."""
    def _do_back():
        _stop_particles()
        if frame_to_hide is not None:
            try:
                frame_to_hide.destroy()
            except Exception:
                pass
        app.minsize(420, 600)
        build_menu(first_time=False)
        
    overlay = ctk.CTkToplevel(app)
    overlay.overrideredirect(True)
    overlay.attributes("-alpha", 0.0)
    overlay.configure(fg_color="#000000")
    overlay.geometry(f"{app.winfo_width()}x{app.winfo_height()}+{app.winfo_rootx()}+{app.winfo_rooty()}")
    overlay.lift()

    def fade_out(step=0, total=12):
        alpha = _smoothstep(step / total)
        try:
            overlay.attributes("-alpha", alpha)
        except Exception:
            return
        if step < total:
            app.after(10, lambda: fade_out(step + 1, total))
        else:
            _do_back()
            fade_in()

    def fade_in(step=0, total=12):
        alpha = 1.0 - _smoothstep(step / total)
        try:
            overlay.attributes("-alpha", alpha)
        except Exception:
            return
        if step < total:
            app.after(10, lambda: fade_in(step + 1, total))
        else:
            try:
                overlay.destroy()
            except Exception:
                pass

    fade_out()


# =============================================
# ЭКРАН НАСТРОЕК
# =============================================

def show_settings_screen():
    _stop_particles()
    # Прячем меню
    for w in app.winfo_children():
        if getattr(w, "is_overlay", False): continue
        try:
            w.place_forget()
            w.pack_forget()
        except Exception:
            pass

    app.minsize(900, 600)

    global settings_screen_frame
    if settings_screen_frame is not None:
        try:
            settings_screen_frame.destroy()
        except Exception:
            pass

    settings_screen_frame = ctk.CTkFrame(app, fg_color="#030303")
    settings_screen_frame.pack(fill="both", expand=True)

    header_frame = ctk.CTkFrame(settings_screen_frame, fg_color="transparent")
    header_frame.pack(fill="x", padx=30, pady=(20, 10))

    ctk.CTkLabel(header_frame, text="⚙ " + T("settings_title"),
                 font=ctk.CTkFont(family="Arial", size=32, weight="bold"),
                 text_color=color_green).pack(side="left")

    main_panel = ctk.CTkFrame(settings_screen_frame, fg_color="transparent",
                               border_width=1, border_color=color_green, corner_radius=16)
    main_panel.pack(fill="both", expand=True, padx=20, pady=(0, 10))

    sidebar = ctk.CTkFrame(main_panel, fg_color="transparent", width=320)
    sidebar.pack(side="left", fill="y", padx=20, pady=30)
    sidebar.pack_propagate(False)

    right_area = ctk.CTkFrame(main_panel, fg_color="transparent",
                               border_width=1, border_color=color_green, corner_radius=12)
    right_area.pack(side="right", fill="both", expand=True, padx=(0, 20), pady=20)

    content_lang    = ctk.CTkFrame(right_area, fg_color="transparent")
    content_reset   = ctk.CTkFrame(right_area, fg_color="transparent")
    content_rewrite = ctk.CTkFrame(right_area, fg_color="transparent")
    content_rules   = ctk.CTkFrame(right_area, fg_color="transparent")

    sidebar_buttons = {}

    def highlight_btn(key):
        for k, b in sidebar_buttons.items():
            if k == key:
                b.configure(fg_color=color_green, text_color="#030303")
            else:
                b.configure(fg_color="transparent", text_color=color_green)

    def show_content(key):
        for c in [content_lang, content_reset, content_rewrite, content_rules]:
            c.pack_forget()
        highlight_btn(key)
        if key == "lang":
            content_lang.pack(fill="both", expand=True, padx=20, pady=20)
        elif key == "reset":
            content_reset.pack(fill="both", expand=True, padx=20, pady=20)
        elif key == "rewrite":
            content_rewrite.pack(fill="both", expand=True, padx=20, pady=20)
        elif key == "rules":
            content_rules.pack(fill="both", expand=True, padx=20, pady=20)

    sidebar_items = [
        ("lang",    "🌐", T("settings_lang")),
        ("reset",   "🗑", T("settings_reset")),
        ("rewrite", "🔄", T("settings_rewrite")),
        ("rules",   "📋", T("settings_rules")),
    ]

    for key, icon, label in sidebar_items:
        btn = ctk.CTkButton(
            sidebar,
            text=f"{icon}  {label}",
            font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
            text_color=color_green,
            fg_color="transparent",
            hover_color="#0a2918",
            anchor="w",
            width=260, height=48,
            corner_radius=10,
            command=lambda k=key: show_content(k)
        )
        btn.pack(pady=6, anchor="w")
        sidebar_buttons[key] = btn

    # ---- LANGUAGE ----
    def set_language(lang_code):
        global current_lang
        current_lang = lang_code
        show_settings_screen()

    lang_code_map = {"en": "lang_english", "ru": "lang_russian", "ky": "lang_kyrgyz"}

    for code, tkey in lang_code_map.items():
        row = ctk.CTkFrame(content_lang, fg_color="transparent")
        row.pack(anchor="w", pady=14, padx=10)
        is_active = (code == current_lang)
        radio = ctk.CTkButton(
            row, text="", width=32, height=32,
            corner_radius=16,
            fg_color=color_green if is_active else "transparent",
            border_color=color_green, border_width=2,
            hover_color="#0a2918",
            command=lambda c=code: set_language(c)
        )
        radio.pack(side="left", padx=(0, 16))
        ctk.CTkLabel(row, text=T(tkey),
                     font=ctk.CTkFont(family="Arial", size=20, weight="bold"),
                     text_color=color_green if is_active else "white").pack(side="left")

    # ---- RESET ----
    ctk.CTkLabel(
        content_reset,
        text=T("reset_warning"),
        font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
        text_color=color_green,
        justify="left",
        wraplength=650
    ).pack(anchor="w", pady=(10, 20))

    btn_row = ctk.CTkFrame(content_reset, fg_color="#0d0d0d",
                            corner_radius=12, border_width=1, border_color="#1a3a25")
    btn_row.pack(fill="x", pady=10, ipady=20, padx=5)

    notif_reset = ctk.CTkLabel(content_reset, text="",
                                font=ctk.CTkFont(family="Arial", size=13),
                                text_color=color_green)

    def do_export_before_reset():
        try:
            from tkinter import filedialog
            import json
            path = filedialog.asksaveasfilename(
                defaultextension=".json",
                initialfile="backup_before_reset.json",
                filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")]
            )
            if not path:
                raise Exception("Отменено пользователем")
            data = {
                "total_games":  StatsData.total_games,
                "success_games": StatsData.success_games,
                "history":      StatsData.history
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            notif_reset.configure(text=f"✓ {path}", text_color=color_green)
            notif_reset.pack(pady=5)
            app.after(3000, notif_reset.pack_forget)
        except Exception as e:
            notif_reset.configure(text=f"✗ {e}", text_color=color_red)
            notif_reset.pack(pady=5)
            app.after(3000, notif_reset.pack_forget)

    def do_reset():
        StatsData.history.clear()
        StatsData.total_games = 0
        StatsData.success_games = 0
        StatsData.success_rate = 0
        StatsData.cyclic_total = 0
        StatsData.cyclic_success = 0
        StatsData.cyclic_fails = 0
        StatsData.random_total = 0
        StatsData.random_success = 0
        StatsData.random_fails = 0
        StatsData.save_db()
        notif_reset.configure(text=T("reset_done"), text_color=color_green)
        notif_reset.pack(pady=5)
        app.after(3000, notif_reset.pack_forget)

    ctk.CTkButton(
        btn_row, text=T("reset_export_btn"),
        font=ctk.CTkFont(family="Arial", size=15, weight="bold"),
        text_color=color_green, fg_color="#0a2918",
        hover_color="#112e1e", corner_radius=20,
        border_width=1, border_color=color_green,
        width=220, height=44,
        command=do_export_before_reset
    ).pack(side="left", padx=20, pady=10)

    ctk.CTkButton(
        btn_row, text=T("reset_btn"),
        font=ctk.CTkFont(family="Arial", size=15, weight="bold"),
        text_color=color_red, fg_color="#2a0a0a",
        hover_color="#3a1010", corner_radius=20,
        border_width=1, border_color=color_red,
        width=220, height=44,
        command=do_reset
    ).pack(side="left", padx=5, pady=10)

    # ---- REWRITE ----
    ctk.CTkLabel(
        content_rewrite,
        text="🔄  " + T("settings_rewrite"),
        font=ctk.CTkFont(family="Arial", size=22, weight="bold"),
        text_color=color_green
    ).pack(anchor="w", pady=(10, 5))

    ctk.CTkLabel(
        content_rewrite,
        text=T("rewrite_desc"),
        font=ctk.CTkFont(family="Arial", size=14),
        text_color="#888888",
        justify="left"
    ).pack(anchor="w", pady=(0, 20))

    def do_rewrite():
        import sys
        python = sys.executable
        os.execl(python, python, *sys.argv)

    ctk.CTkButton(
        content_rewrite,
        text="🔄  " + T("settings_rewrite"),
        font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
        text_color=color_green, fg_color="#0a2918",
        hover_color="#112e1e", corner_radius=20,
        border_width=1, border_color=color_green,
        width=240, height=48,
        command=do_rewrite
    ).pack(anchor="w")

    # ---- RULES ----
    scroll_rules = ctk.CTkScrollableFrame(content_rules, fg_color="transparent")
    scroll_rules.pack(fill="both", expand=True)

    ctk.CTkLabel(
        scroll_rules,
        text=T("rules_text"),
        font=ctk.CTkFont(family="Arial", size=14),
        text_color=color_green,
        justify="left",
        wraplength=680,
        anchor="nw"
    ).pack(anchor="nw", padx=10, pady=10)

    ctk.CTkButton(
        settings_screen_frame,
        text=T("btn_exit"),
        font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
        text_color=color_green,
        fg_color="transparent",
        hover_color="#0a2918",
        width=120, height=40,
        command=lambda: back_to_menu(settings_screen_frame)
    ).pack(side="bottom", anchor="se", padx=30, pady=15)

    show_content("lang")


# =============================================
# ЭКРАН ЭКСПОРТА
# =============================================

def show_export_screen():
    _stop_particles()
    for w in app.winfo_children():
        if getattr(w, "is_overlay", False): continue
        try:
            w.place_forget()
            w.pack_forget()
        except Exception:
            pass

    app.minsize(900, 600)

    global export_screen_frame
    if export_screen_frame is not None:
        try:
            export_screen_frame.destroy()
        except Exception:
            pass

    export_screen_frame = ctk.CTkFrame(app, fg_color="#030303")
    export_screen_frame.pack(fill="both", expand=True)

    header_frame = ctk.CTkFrame(export_screen_frame, fg_color="transparent")
    header_frame.pack(fill="x", padx=30, pady=(25, 5))

    ctk.CTkLabel(header_frame, text="🗄",
                 font=ctk.CTkFont(size=36), text_color=color_green).pack(side="left", padx=(0, 10))

    ctk.CTkLabel(header_frame, text=T("export_title"),
                 font=ctk.CTkFont(family="Arial", size=32, weight="bold"),
                 text_color=color_green).pack(side="left")

    ctk.CTkLabel(export_screen_frame,
                 text=T("export_sub"),
                 font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                 text_color=color_green).pack(anchor="w", padx=30, pady=(0, 15))

    scroll = ctk.CTkScrollableFrame(export_screen_frame, fg_color="transparent")
    scroll.pack(fill="both", expand=True, padx=20, pady=(0, 10))

    notif_lbl = ctk.CTkLabel(scroll, text="",
                              font=ctk.CTkFont(family="Arial", size=13),
                              text_color=color_green)

    def show_notif(msg, success=True):
        notif_lbl.configure(text=msg, text_color=color_green if success else color_red)
        notif_lbl.pack(pady=5)
        app.after(3000, lambda: notif_lbl.pack_forget())

    def get_save_path(filename):
        from tkinter import filedialog
        import os
        ext = os.path.splitext(filename)[1]
        ext_desc = "JSON Files" if ext == ".json" else "CSV Files" if ext == ".csv" else "Excel Files"
        path = filedialog.asksaveasfilename(
            defaultextension=ext,
            initialfile=filename,
            filetypes=[(ext_desc, f"*{ext}"), ("All Files", "*.*")]
        )
        if not path:
            raise Exception("Отменено пользователем / Cancelled by user")
        return path

    def make_export_card(parent, icon_text, title, description, csv_fn, excel_fn, json_fn):
        card = ctk.CTkFrame(parent, fg_color="#0d0d0d",
                             corner_radius=16, border_width=1, border_color="#1a3a25")
        card.pack(fill="x", padx=10, pady=10, ipady=15)

        top_row = ctk.CTkFrame(card, fg_color="transparent")
        top_row.pack(fill="x", padx=20, pady=(15, 5))

        icon_box = ctk.CTkFrame(top_row, width=60, height=60,
                                fg_color="#0a2918", corner_radius=12)
        icon_box.pack(side="left", padx=(0, 15))
        icon_box.pack_propagate(False)
        ctk.CTkLabel(icon_box, text=icon_text,
                     font=ctk.CTkFont(size=30)).place(relx=0.5, rely=0.5, anchor="center")

        text_block = ctk.CTkFrame(top_row, fg_color="transparent")
        text_block.pack(side="left", fill="both", expand=True)

        ctk.CTkLabel(text_block, text=title,
                     font=ctk.CTkFont(family="Arial", size=20, weight="bold"),
                     text_color=color_green, anchor="w").pack(anchor="w")
        ctk.CTkLabel(text_block, text=description,
                     font=ctk.CTkFont(family="Arial", size=12),
                     text_color="#888888", anchor="w",
                     wraplength=700, justify="left").pack(anchor="w")

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(anchor="w", padx=20, pady=(10, 5))

        for label, icon, fn in [
            (T("btn_csv"),   "☁", csv_fn),
            (T("btn_excel"), "✕", excel_fn),
            (T("btn_json"),  "📄", json_fn),
        ]:
            ctk.CTkButton(
                btn_row, text=f"{icon}  {label}",
                font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                text_color=color_green, fg_color="#0a2918",
                hover_color="#112e1e", corner_radius=20,
                border_width=1, border_color=color_green,
                width=220, height=42, command=fn
            ).pack(side="left", padx=(0, 12))

    # --- Export functions ---
    def export_stats_csv():
        try:
            path = get_save_path("statistics.csv")
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Показатель", "Значение"])
                w.writerow([T("stats_total"),   StatsData.total_games])
                w.writerow([T("stats_success"), StatsData.success_games])
                w.writerow([T("stats_rate"),    f"{StatsData.success_rate}%"])
                w.writerow([])
                w.writerow([T("stats_cyclic") + " — " + T("stats_total_g"), StatsData.cyclic_total])
                w.writerow([T("stats_cyclic") + " — " + T("stats_succ"),    StatsData.cyclic_success])
                w.writerow([T("stats_cyclic") + " — " + T("stats_fail"),    StatsData.cyclic_fails])
                w.writerow([])
                w.writerow([T("stats_random") + " — " + T("stats_total_g"), StatsData.random_total])
                w.writerow([T("stats_random") + " — " + T("stats_succ"),    StatsData.random_success])
                w.writerow([T("stats_random") + " — " + T("stats_fail"),    StatsData.random_fails])
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_stats_excel():
        try:
            try:
                import openpyxl
                path = get_save_path("statistics.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = T("stats_tab_stats")
                ws.append([T("stats_total"),   StatsData.total_games])
                ws.append([T("stats_success"), StatsData.success_games])
                ws.append([T("stats_rate"),    f"{StatsData.success_rate}%"])
                wb.save(path)
            except ImportError:
                path = get_save_path("statistics.xls")
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    csv.writer(f, delimiter="\t").writerow([T("stats_total"), StatsData.total_games])
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_stats_json():
        try:
            path = get_save_path("statistics.json")
            data = {
                "total_games":   StatsData.total_games,
                "success_games": StatsData.success_games,
                "success_rate":  StatsData.success_rate,
                "cyclic":  {"total": StatsData.cyclic_total,  "success": StatsData.cyclic_success,  "fails": StatsData.cyclic_fails},
                "random":  {"total": StatsData.random_total,  "success": StatsData.random_success,  "fails": StatsData.random_fails}
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_history_csv():
        try:
            path = get_save_path("history.csv")
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["№", T("history_win") + "/" + T("history_lose"), "Strategy", "%", T("history_date"), T("history_time")])
                for i, item in enumerate(StatsData.history, 1):
                    w.writerow([i, T("history_win") if item['is_win'] else T("history_lose"),
                                item['strategy'], f"{item['rate']}%", item['date'], item['time']])
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_history_excel():
        try:
            try:
                import openpyxl
                path = get_save_path("history.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = T("stats_tab_history")
                ws.append(["№", "Result", "Strategy", "%", "Date", "Time"])
                for i, item in enumerate(StatsData.history, 1):
                    ws.append([i, T("history_win") if item['is_win'] else T("history_lose"),
                               item['strategy'], f"{item['rate']}%", item['date'], item['time']])
                wb.save(path)
            except ImportError:
                path = get_save_path("history.xls")
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    pass
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_history_json():
        try:
            path = get_save_path("history.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump(StatsData.history, f, ensure_ascii=False, indent=2)
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_detail_csv():
        try:
            path = get_save_path("detailed_history.csv")
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["#", "Result", "Strategy", "%", "Date", "Time"])
                for i, item in enumerate(StatsData.history, 1):
                    w.writerow([i, T("history_win") if item['is_win'] else T("history_lose"),
                                item['strategy'], f"{item['rate']}%", item['date'], item['time']])
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_detail_excel():
        try:
            try:
                import openpyxl
                path = get_save_path("detailed_history.xlsx")
                wb = openpyxl.Workbook()
                for i, item in enumerate(StatsData.history, 1):
                    ws = wb.create_sheet(title=f"Game {i}")
                    ws.append(["Result",   T("history_win") if item['is_win'] else T("history_lose")])
                    ws.append(["Strategy", item['strategy']])
                    ws.append(["Rate",     f"{item['rate']}%"])
                    ws.append(["Date",     item['date']])
                    ws.append(["Time",     item['time']])
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                if not wb.sheetnames:
                    wb.create_sheet("No data")
                wb.save(path)
            except ImportError:
                path = get_save_path("detailed_history.xls")
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    pass
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    def export_detail_json():
        try:
            path = get_save_path("detailed_history.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump([{"game": i, **item} for i, item in enumerate(StatsData.history, 1)],
                          f, ensure_ascii=False, indent=2)
            show_notif(f"✓ {path}")
        except Exception as e:
            show_notif(f"✗ {e}", False)

    make_export_card(scroll, "📊", T("export_stats_title"), T("export_stats_desc"),
                     export_stats_csv, export_stats_excel, export_stats_json)
    make_export_card(scroll, "🗂", T("export_history_title"), T("export_history_desc"),
                     export_history_csv, export_history_excel, export_history_json)
    make_export_card(scroll, "🗄", T("export_detail_title"), T("export_detail_desc"),
                     export_detail_csv, export_detail_excel, export_detail_json)

    notif_lbl.pack(pady=5)
    notif_lbl.pack_forget()

    exit_btn_ex = ctk.CTkButton(
        export_screen_frame,
        text=T("btn_exit"),
        font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
        text_color=color_green, fg_color="transparent",
        hover_color="#0a2918", width=120, height=40,
        command=lambda: back_to_menu(export_screen_frame)
    )
    exit_btn_ex.pack(side="bottom", anchor="se", padx=30, pady=15)
    exit_btn_ex.bind("<Enter>", lambda e: exit_btn_ex.configure(text_color="#FF0000"))
    exit_btn_ex.bind("<Leave>", lambda e: exit_btn_ex.configure(text_color=color_green))
    exit_btn_ex.bind("<ButtonPress-1>", lambda e: exit_btn_ex.configure(text_color="#FF0000"))


# =============================================
# ИГРОВОЙ ЭКРАН
# =============================================

def create_rounded_gradient(width, height, radius, color_top, color_bot):
    from PIL import Image, ImageDraw
    gradient = Image.new("RGB", (width, height))
    draw = ImageDraw.Draw(gradient)
    for y in range(height):
        r = int(color_top[0] + (color_bot[0] - color_top[0]) * y / height)
        g = int(color_top[1] + (color_bot[1] - color_top[1]) * y / height)
        b = int(color_top[2] + (color_bot[2] - color_top[2]) * y / height)
        draw.line([(0, y), (width, y)], fill=(r, g, b))
    return gradient

def show_game_screen():
    _stop_particles()
    for w in app.winfo_children():
        if getattr(w, "is_overlay", False): continue
        try:
            w.place_forget()
            w.pack_forget()
        except Exception:
            pass

    app.minsize(900, 600)

    global game_screen
    if game_screen is not None:
        try:
            game_screen.destroy()
        except Exception:
            pass

    game_screen = ctk.CTkFrame(app, fg_color="#030303")
    game_screen.pack(fill="both", expand=True)

    main_container = ctk.CTkFrame(game_screen, fg_color="transparent")
    main_container.place(relx=0.5, rely=0.5, anchor="center")

    square_size = 200
    canvas_frame = ctk.CTkFrame(main_container, width=square_size, height=square_size, corner_radius=20, fg_color="transparent")
    canvas_frame.pack(pady=(0, 20))
    canvas_frame.pack_propagate(False)

    color_top = (44, 228, 107)    # #2ce46b
    color_bot = (3, 83, 35)       # #035323
    grad_img = create_rounded_gradient(square_size, square_size, 25, color_top, color_bot)
    grad_img = grad_img.convert("RGBA")

    dice_path = r"C:\Users\user\.gemini\antigravity\brain\c5648e94-ebe2-4867-bf3b-b2758ec1f675\die_icon_red_dot_1779024620989.png"
    try:
        from PIL import Image, ImageDraw
        dice_img = Image.open(dice_path).convert("RGBA")
        bbox = dice_img.getbbox()
        if bbox:
            dice_img = dice_img.crop(bbox)
        ratio = max(square_size / dice_img.width, square_size / dice_img.height)
        new_w = int(dice_img.width * ratio)
        new_h = int(dice_img.height * ratio)
        dice_img = dice_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        x = (square_size - new_w) // 2
        y = (square_size - new_h) // 2
        grad_img.paste(dice_img, (x, y), dice_img)
        mask = Image.new("L", (square_size, square_size), 0)
        draw_mask = ImageDraw.Draw(mask)
        draw_mask.rounded_rectangle((0, 0, square_size, square_size), radius=25, fill=255)
        final_output = Image.new("RGBA", (square_size, square_size), (0, 0, 0, 0))
        final_output.paste(grad_img, (0, 0), mask=mask)
        grad_img = final_output
    except Exception as e:
        print(f"Не удалось загрузить иконку кубика: {e}")

    menu_grad_img = ctk.CTkImage(light_image=grad_img, dark_image=grad_img, size=(square_size, square_size))
    icon_label = ctk.CTkLabel(canvas_frame, text="", image=menu_grad_img)
    icon_label.pack(fill="both", expand=True)

    ctk.CTkLabel(main_container, text="100", font=("Arial", 65, "bold"), text_color="#12A850").pack(pady=(10, 0))
    ctk.CTkLabel(main_container, text=T("title_word"), font=("Arial", 55, "bold"), text_color="#12A850").pack(pady=(0, 10))
    
    ctk.CTkLabel(main_container, text=T("game_subtitle").replace('\n', ' '), font=("Arial", 22, "bold"), text_color="#DDDDDD").pack(pady=(10, 5))
    ctk.CTkLabel(main_container, text=T("game_desc").replace('\n', ' '), font=("Arial", 14), text_color="#888888", justify="center").pack(pady=(0, 40))

    btns_frame = ctk.CTkFrame(main_container, fg_color="transparent")
    btns_frame.pack()

    btn_manual = ctk.CTkButton(btns_frame, text=T("btn_manual"), font=("Arial", 22, "bold"), height=60, width=280, 
                               corner_radius=30, fg_color="#12A850", hover_color="#2CE46B", text_color="#FFFFFF",
                               command=lambda: transition_to(show_manual_sim_screen))
    btn_manual.pack(side="left", padx=10)
    
    btn_auto = ctk.CTkButton(btns_frame, text=T("btn_auto"), font=("Arial", 22, "bold"), height=60, width=280, 
                             corner_radius=30, fg_color="#12A850", hover_color="#2CE46B", text_color="#FFFFFF",
                             command=lambda: transition_to(show_auto_sim_screen))
    btn_auto.pack(side="left", padx=10)

    ctk.CTkButton(
        main_container, text=T("back_to_menu"),
        font=ctk.CTkFont(family="Arial", size=14),
        fg_color="transparent", text_color="#909090",
        hover_color="#1a1a1a", width=150, height=40,
        command=lambda: back_to_menu(game_screen)
    ).pack(pady=20)


# =============================================
# РУЧНАЯ ИГРА
# =============================================

def show_manual_sim_screen():
    if game_screen is not None:
        try:
            game_screen.pack_forget()
        except Exception:
            pass
    app.minsize(900, 600)

    global manual_sim_frame
    if manual_sim_frame is not None:
        try:
            manual_sim_frame.destroy()
        except Exception:
            pass

    manual_sim_frame = ctk.CTkFrame(app, fg_color="#030303")
    manual_sim_frame.pack(fill="both", expand=True)

    left_frame = ctk.CTkFrame(manual_sim_frame, fg_color="transparent", width=350)
    left_frame.pack(side="left", fill="y", padx=40, pady=40)

    right_frame = ctk.CTkFrame(manual_sim_frame, fg_color="transparent")
    right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=40)

    # Left Column Container
    left_inner = ctk.CTkFrame(left_frame, fg_color="transparent")
    left_inner.pack(anchor="n")

    square_size = 180
    canvas_frame = ctk.CTkFrame(left_inner, width=square_size, height=square_size, corner_radius=25, fg_color="transparent")
    canvas_frame.pack(pady=(0, 10))
    canvas_frame.pack_propagate(False)

    color_top = (44, 228, 107)
    color_bot = (3, 83, 35)
    grad_img = create_rounded_gradient(square_size, square_size, 25, color_top, color_bot)
    grad_img = grad_img.convert("RGBA")

    dice_path = r"C:\Users\user\.gemini\antigravity\brain\c5648e94-ebe2-4867-bf3b-b2758ec1f675\die_icon_red_dot_1779024620989.png"
    try:
        from PIL import Image, ImageDraw
        dice_img = Image.open(dice_path).convert("RGBA")
        bbox = dice_img.getbbox()
        if bbox:
            dice_img = dice_img.crop(bbox)
        ratio = max(square_size / dice_img.width, square_size / dice_img.height)
        new_w = int(dice_img.width * ratio)
        new_h = int(dice_img.height * ratio)
        dice_img = dice_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        x = (square_size - new_w) // 2
        y = (square_size - new_h) // 2
        grad_img.paste(dice_img, (x, y), dice_img)
        mask = Image.new("L", (square_size, square_size), 0)
        draw_mask = ImageDraw.Draw(mask)
        draw_mask.rounded_rectangle((0, 0, square_size, square_size), radius=25, fill=255)
        final_output = Image.new("RGBA", (square_size, square_size), (0, 0, 0, 0))
        final_output.paste(grad_img, (0, 0), mask=mask)
        grad_img = final_output
    except Exception as e:
        pass

    menu_grad_img = ctk.CTkImage(light_image=grad_img, dark_image=grad_img, size=(square_size, square_size))
    icon_label = ctk.CTkLabel(canvas_frame, text="", image=menu_grad_img)
    icon_label.pack(fill="both", expand=True)

    ctk.CTkLabel(left_inner, text="100",
                 font=ctk.CTkFont(family="Arial Black", size=40, weight="bold"),
                 text_color="white").pack()
    ctk.CTkLabel(left_inner, text=T("title_word"),
                 font=ctk.CTkFont(family="Arial Black", size=30, weight="bold"),
                 text_color="white").pack(pady=(0, 20))

    status_frame = ctk.CTkFrame(left_inner, fg_color="#1eb852", corner_radius=20, width=280, height=90)
    status_frame.pack(pady=10)
    status_frame.pack_propagate(False)
    
    lbl_current_prisoner = ctk.CTkLabel(status_frame,
                                        text=T("prisoner_label") + "1",
                                        font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                                        text_color="white")
    lbl_current_prisoner.pack(anchor="w", padx=20, pady=(15, 5))
    lbl_attempts_left = ctk.CTkLabel(status_frame,
                                     text=T("tries_left").replace("\n", " ") + " 50",
                                     font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                                     text_color="white", justify="left")
    lbl_attempts_left.pack(anchor="w", padx=20)

    stats_row = ctk.CTkFrame(left_inner, fg_color="transparent")
    stats_row.pack(pady=10)

    s_frame1 = ctk.CTkFrame(stats_row, fg_color="#1eb852", corner_radius=15, width=135, height=80)
    s_frame1.pack(side="left", padx=(0, 5))
    s_frame1.pack_propagate(False)
    ctk.CTkLabel(s_frame1, text=T("succ_prisoners").replace("\n", " "),
                 font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                 text_color="white", justify="left").pack(anchor="w", padx=10, pady=(10, 0))
    lbl_success_count = ctk.CTkLabel(s_frame1, text="#0/100",
                                     font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                                     text_color="white")
    lbl_success_count.pack(anchor="w", padx=10, pady=(2, 0))

    s_frame2 = ctk.CTkFrame(stats_row, fg_color="#1eb852", corner_radius=15, width=135, height=80)
    s_frame2.pack(side="left", padx=(5, 0))
    s_frame2.pack_propagate(False)
    ctk.CTkLabel(s_frame2, text=T("used_tries").replace("\n", " "),
                 font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                 text_color="white", justify="left").pack(anchor="w", padx=10, pady=(10, 0))
    lbl_attempts_used = ctk.CTkLabel(s_frame2, text="#0/50",
                                     font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                                     text_color="white")
    lbl_attempts_used.pack(anchor="w", padx=10, pady=(2, 0))

    exit_btn = ctk.CTkButton(left_frame, text=T("btn_exit"),
                  font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
                  fg_color="transparent", text_color="#1eb852", hover_color="#030303",
                  command=lambda: back_to_game_screen_from_auto(manual_sim_frame))
    exit_btn.pack(side="bottom", anchor="sw", pady=20)
    
    exit_btn.bind("<Enter>", lambda e: exit_btn.configure(text_color="#FF0000"))
    exit_btn.bind("<Leave>", lambda e: exit_btn.configure(text_color="#1eb852"))
    exit_btn.bind("<ButtonPress-1>", lambda e: exit_btn.configure(text_color="#FF0000"))

    grid_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
    grid_frame.place(relx=0.5, rely=0.5, anchor="center")

    boxes = list(range(1, 101))
    import random
    random.shuffle(boxes)
    current_prisoner = 1
    attempts_used = 0
    successes = 0
    opened_boxes = set()
    game_over = False

    def update_labels():
        lbl_current_prisoner.configure(text=T("prisoner_label") + str(current_prisoner))
        lbl_attempts_left.configure(text=T("tries_left").replace("\n", " ") + " " + str(50 - attempts_used))
        lbl_success_count.configure(text=f"#{successes}/100")
        lbl_attempts_used.configure(text=f"#{attempts_used}/50")

    def box_clicked(index):
        nonlocal current_prisoner, attempts_used, successes, game_over

        if game_over or index in opened_boxes:
            return

        opened_boxes.add(index)
        attempts_used += 1
        revealed_number = boxes[index]
        box_buttons[index].configure(text=str(revealed_number), fg_color="#808080", hover_color="#808080")
        update_labels()

        if revealed_number == current_prisoner:
            successes += 1
            update_labels()
            for b in box_buttons:
                b.configure(state="disabled")

            if successes == 100:
                game_over = True
                lbl_current_prisoner.configure(text=T("game_won"))
                StatsData.total_games += 1
                StatsData.random_total += 1
                StatsData.random_success += 1
                StatsData.save_db()
                StatsData.success_games += 1
                if StatsData.total_games > 0:
                    StatsData.success_rate = int((StatsData.success_games / StatsData.total_games) * 100)
                import datetime
                now = datetime.datetime.now()
                StatsData.history.insert(0, {
                    "is_win":   True,
                    "strategy": "Ручная игра",
                    "rate":     100,
                    "date":     now.strftime("%d.%m.%y"),
                    "time":     now.strftime("%H:%M:%S")
                })
                return

            app.after(1000, next_prisoner)

        elif attempts_used >= 50:
            game_over = True
            lbl_current_prisoner.configure(text=T("game_lost"))
            for b in box_buttons:
                b.configure(state="disabled")

            StatsData.total_games += 1
            StatsData.random_total += 1
            StatsData.random_fails += 1
            StatsData.save_db()
            if StatsData.total_games > 0:
                StatsData.success_rate = int((StatsData.success_games / StatsData.total_games) * 100)
            import datetime
            now = datetime.datetime.now()
            StatsData.history.insert(0, {
                "is_win":   False,
                "strategy": "Ручная игра",
                "rate":     int((successes / 100) * 100),
                "date":     now.strftime("%d.%m.%y"),
                "time":     now.strftime("%H:%M:%S")
            })

    def next_prisoner():
        nonlocal current_prisoner, attempts_used
        current_prisoner += 1
        attempts_used = 0
        opened_boxes.clear()
        for i, btn in enumerate(box_buttons):
            btn.configure(text=str(i + 1), fg_color="#1eb852", hover_color="#11a64a", state="normal")
        update_labels()

    box_buttons = []
    for i in range(100):
        row = i // 10
        col = i % 10
        btn = ctk.CTkButton(grid_frame, text=str(i + 1), width=50, height=50,
                             fg_color="#1eb852", text_color="white",
                             hover_color="#11a64a",
                             font=ctk.CTkFont(family="Arial", size=22, weight="bold"),
                             corner_radius=10)
        btn.grid(row=row, column=col, padx=4, pady=4)
        btn.configure(command=lambda idx=i: box_clicked(idx))
        btn.is_cell = True
        box_buttons.append(btn)

    update_labels()

# =============================================
# АВТО СИМУЛЯЦИЯ
# =============================================

def show_auto_sim_screen():
    if game_screen is not None:
        try:
            game_screen.pack_forget()
        except Exception:
            pass
    app.minsize(1000, 700)

    global auto_sim_frame
    if auto_sim_frame is not None:
        try:
            auto_sim_frame.destroy()
        except Exception:
            pass

    auto_sim_frame = ctk.CTkFrame(app, fg_color="#111111")
    auto_sim_frame.pack(fill="both", expand=True)

    session_total_games = 0
    session_total_wins = 0
    session_total_losses = 0

    selected_prisoners = 100
    selected_games = 100
    selected_strategy = "cycles"
    box_buttons = []
    is_simulating = False

    # --- ВЕРХНЯЯ ЧАСТЬ (TOP SECTION) ---
    top_section = ctk.CTkFrame(auto_sim_frame, fg_color="transparent")
    top_section.pack(fill="x", padx=20, pady=(20, 10))

    left_col = ctk.CTkFrame(top_section, fg_color="transparent")
    left_col.pack(side="left", fill="both", expand=True, padx=(0, 10))

    header_frame = ctk.CTkFrame(left_col, fg_color="transparent")
    header_frame.pack(fill="x")

    icon_frame = ctk.CTkFrame(header_frame, width=90, height=90, fg_color="#1D8B4E", corner_radius=15)
    icon_frame.pack(side="left", padx=(0, 20))
    icon_frame.pack_propagate(False)

    image_path = r"C:\Users\user\.gemini\antigravity\brain\c5648e94-ebe2-4867-bf3b-b2758ec1f675\die_icon_red_dot_1779024620989.png"
    try:
        from PIL import Image
        die_img = ctk.CTkImage(light_image=Image.open(image_path), dark_image=Image.open(image_path), size=(80, 80))
        icon_label = ctk.CTkLabel(icon_frame, text="", image=die_img)
    except Exception:
        icon_label = ctk.CTkLabel(icon_frame, text="🎲", font=("Arial", 45))

    icon_label.place(relx=0.5, rely=0.5, anchor="center")

    titles_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
    titles_frame.pack(side="left", fill="both", expand=True)
    lbl_title = ctk.CTkLabel(titles_frame, text=T("auto_title"), font=("Arial", 18, "bold"), text_color="#FFFFFF")
    lbl_title.pack(anchor="w", pady=(15, 5))
    lbl_sub = ctk.CTkLabel(titles_frame, text=T("auto_sub"), font=("Arial", 11, "bold"), text_color="#FFFFFF")
    lbl_sub.pack(anchor="w")

    settings_frame = ctk.CTkFrame(left_col, fg_color="#1A1A1A", corner_radius=15, border_width=1, border_color="#333333")
    settings_frame.pack(fill="x", pady=(20, 0))

    lbl_set_title = ctk.CTkLabel(settings_frame, text=T("sim_settings"), font=("Arial", 14, "bold"), text_color="#1D8B4E")
    lbl_set_title.pack(anchor="w", padx=15, pady=(10, 5))

    set_controls = ctk.CTkFrame(settings_frame, fg_color="transparent")
    set_controls.pack(fill="x", padx=15, pady=(0, 15))

    p_frame = ctk.CTkFrame(set_controls, fg_color="transparent")
    p_frame.pack(side="left")
    ctk.CTkLabel(p_frame, text=T("num_prisoners"), font=("Arial", 12, "bold"), text_color="#FFFFFF").pack(anchor="w", pady=(0, 5))
    p_btns_frame = ctk.CTkFrame(p_frame, fg_color="transparent")
    p_btns_frame.pack(anchor="w")

    def set_prisoners(val):
        nonlocal selected_prisoners
        selected_prisoners = val
        active = "#1D8B4E"
        inactive = "#555555"
        for btn, v in [(btn_p10, 10), (btn_p50, 50), (btn_p100, 100)]:
            color = active if v == val else inactive
            btn.configure(fg_color=color, hover_color=color)
        create_grid()

    btn_p10 = ctk.CTkButton(p_btns_frame, text="10", width=60, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_prisoners(10))
    btn_p10.pack(side="left", padx=(0, 5))
    btn_p50 = ctk.CTkButton(p_btns_frame, text="50", width=60, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_prisoners(50))
    btn_p50.pack(side="left", padx=(0, 5))
    btn_p100 = ctk.CTkButton(p_btns_frame, text="100", width=60, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_prisoners(100))
    btn_p100.pack(side="left")

    g_frame = ctk.CTkFrame(set_controls, fg_color="transparent")
    g_frame.pack(side="right")
    ctk.CTkLabel(g_frame, text=T("num_games"), font=("Arial", 12, "bold"), text_color="#FFFFFF").pack(anchor="w", pady=(0, 5))
    g_btns_frame = ctk.CTkFrame(g_frame, fg_color="transparent")
    g_btns_frame.pack(anchor="w")

    def set_games(val):
        nonlocal selected_games
        selected_games = val
        active = "#1D8B4E"
        inactive = "#555555"
        for btn, v in [(btn_g100, 100), (btn_g500, 500), (btn_g1000, 1000)]:
            color = active if v == val else inactive
            btn.configure(fg_color=color, hover_color=color)

    btn_g100 = ctk.CTkButton(g_btns_frame, text="100", width=70, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_games(100))
    btn_g100.pack(side="left", padx=(0, 5))
    btn_g500 = ctk.CTkButton(g_btns_frame, text="500", width=70, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_games(500))
    btn_g500.pack(side="left", padx=(0, 5))
    btn_g1000 = ctk.CTkButton(g_btns_frame, text="1000", width=70, font=("Arial", 14, "bold"), corner_radius=5, text_color="#FFFFFF", command=lambda: set_games(1000))
    btn_g1000.pack(side="left")

    right_col = ctk.CTkFrame(top_section, fg_color="#1A1A1A", corner_radius=15, border_width=1, border_color="#333333")
    right_col.pack(side="right", fill="both", expand=True, padx=(10, 0))

    lbl_rules_title = ctk.CTkLabel(right_col, text=T("rules_title_auto"), font=("Arial", 16, "bold"), text_color="#1D8B4E")
    lbl_rules_title.pack(anchor="w", padx=15, pady=(15, 5))

    lbl_rules = ctk.CTkLabel(right_col, text=T("rules_auto"), font=("Arial", 14), text_color="#888888", justify="left")
    lbl_rules.pack(anchor="w", padx=15, pady=(0, 15))

    # --- СРЕДНЯЯ ЧАСТЬ (MIDDLE SECTION) ---
    mid_frame = ctk.CTkFrame(auto_sim_frame, fg_color="#1A1A1A", corner_radius=15, border_width=1, border_color="#333333")
    mid_frame.pack(fill="x", padx=20, pady=10)

    lbl_strat_title = ctk.CTkLabel(mid_frame, text=T("choose_strategy"), font=("Arial", 22, "bold"), text_color="#FFFFFF")
    lbl_strat_title.place(x=25, y=10)

    strat_btn_frame = ctk.CTkFrame(mid_frame, fg_color="transparent")
    strat_btn_frame.pack(side="left", fill="x", expand=True, padx=(20, 10), pady=(50, 20))

    def set_strategy(strat):
        nonlocal selected_strategy
        selected_strategy = strat
        active = "#1eb852"
        inactive = "#555555"
        if strat == "cycles":
            btn_strat_cycles.configure(fg_color=active, hover_color=active)
            btn_strat_random.configure(fg_color=inactive, hover_color=inactive)
        else:
            btn_strat_cycles.configure(fg_color=inactive, hover_color=inactive)
            btn_strat_random.configure(fg_color=active, hover_color=active)

    btn_strat_cycles = ctk.CTkButton(strat_btn_frame, text=T("strat_cyclic"), font=("Arial", 30, "bold"), height=75, corner_radius=10, text_color="#FFFFFF", command=lambda: set_strategy("cycles"))
    btn_strat_cycles.pack(side="left", padx=(0, 15), expand=True, fill="x")

    btn_strat_random = ctk.CTkButton(strat_btn_frame, text=T("strat_random"), font=("Arial", 30, "bold"), height=75, corner_radius=10, text_color="#FFFFFF", command=lambda: set_strategy("random"))
    btn_strat_random.pack(side="left", padx=(15, 0), expand=True, fill="x")

    start_frame = ctk.CTkFrame(mid_frame, fg_color="transparent")
    start_frame.pack(side="right", padx=25, pady=(15, 20))

    def run_simulation():
        nonlocal is_simulating
        if is_simulating:
            return
        is_simulating = True
        btn_start.configure(state="disabled")
        for btn in box_buttons:
            btn.configure(fg_color="#1D8B4E", hover_color="#1D8B4E")
            
        import threading
        import random
        import datetime
        
        def worker():
            p_count = selected_prisoners
            g_count = selected_games
            strat = selected_strategy
            
            half = p_count // 2
            wins = 0
            last_opened = []
            
            for g in range(g_count):
                boxes = list(range(1, p_count + 1))
                random.shuffle(boxes)
                game_win = True
                
                for p in range(1, p_count + 1):
                    opened_boxes = []
                    if strat == "cycles":
                        current_box = p
                        found = False
                        for step in range(half):
                            opened_boxes.append(current_box)
                            content = boxes[current_box - 1]
                            if content == p:
                                found = True
                                break
                            current_box = content
                    else:
                        opened_boxes = random.sample(range(1, p_count + 1), half)
                        found = False
                        for box_idx in opened_boxes:
                            content = boxes[box_idx - 1]
                            if content == p:
                                found = True
                                break
                                
                    if g == g_count - 1:
                        last_opened = opened_boxes
                        
                    if not found:
                        game_win = False
                        if g != g_count - 1:
                            break
                            
                if game_win:
                    wins += 1
                    
            app.after(0, update_stats, g_count, wins, last_opened, strat)

        threading.Thread(target=worker).start()

    btn_start = ctk.CTkButton(start_frame, text="▶", font=("Arial", 40), width=80, height=70, fg_color="#1D8B4E", hover_color="#187540", text_color="#FFFFFF", corner_radius=15, command=run_simulation)
    btn_start.pack()
    ctk.CTkLabel(start_frame, text=T("start_btn"), font=("Arial", 18, "bold"), text_color="#1D8B4E").pack(pady=(5, 0))

    # --- НИЖНЯЯ ЧАСТЬ (BOTTOM SECTION) ---
    bot_frame = ctk.CTkFrame(auto_sim_frame, fg_color="transparent")
    bot_frame.pack(fill="both", expand=True, padx=20, pady=(10, 20))

    grid_frame = ctk.CTkFrame(bot_frame, fg_color="#1A1A1A", corner_radius=15, border_width=1, border_color="#333333")
    grid_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

    stats_outer = ctk.CTkFrame(bot_frame, fg_color="transparent")
    stats_outer.pack(side="right", fill="y", padx=(10, 0))

    stats_grid = ctk.CTkFrame(stats_outer, fg_color="transparent")
    stats_grid.pack(anchor="n")

    def create_stat_box(parent, text, row, col):
        btn = ctk.CTkButton(parent, text=text, font=("Arial", 12, "bold"), fg_color="#1D8B4E", hover_color="#1D8B4E", text_color="#FFFFFF", width=95, height=75, corner_radius=10)
        btn.grid(row=row, column=col, padx=5, pady=5)
        return btn

    lbl_total_games = create_stat_box(stats_grid, f"{T('total_games_res')}\n0", 0, 0)
    lbl_success_rate = create_stat_box(stats_grid, f"{T('success_pct')}\n0%", 0, 1)
    lbl_wins = create_stat_box(stats_grid, f"{T('success_res')}\n0", 1, 0)
    lbl_losses = create_stat_box(stats_grid, f"{T('fail_res')}\n0", 1, 1)

    exit_btn = ctk.CTkButton(stats_outer, text=T("btn_exit"), font=("Arial", 24, "bold"), fg_color="transparent", text_color="#1eb852", hover_color="#111111", command=lambda: back_to_game_screen_from_auto(auto_sim_frame))
    exit_btn.pack(side="bottom", anchor="e", pady=10)

    exit_btn.bind("<Enter>", lambda e: exit_btn.configure(text_color="#FF0000"))
    exit_btn.bind("<Leave>", lambda e: exit_btn.configure(text_color="#1eb852"))
    exit_btn.bind("<ButtonPress-1>", lambda e: exit_btn.configure(text_color="#FF0000"))

    def create_grid():
        for widget in grid_frame.winfo_children():
            widget.destroy()
            
        nonlocal box_buttons
        box_buttons = []
        cols = 20 if selected_prisoners >= 20 else selected_prisoners
        rows = (selected_prisoners + cols - 1) // cols
        
        for i in range(20):
            grid_frame.grid_columnconfigure(i, weight=0)
            grid_frame.grid_rowconfigure(i, weight=0)
            
        for i in range(cols):
            grid_frame.grid_columnconfigure(i, weight=1)
        for i in range(rows):
            grid_frame.grid_rowconfigure(i, weight=1)
            
        for i in range(selected_prisoners):
            btn = ctk.CTkButton(grid_frame, text=str(i+1), font=("Arial", 26, "bold"), width=10, height=10, border_spacing=1, fg_color="#1D8B4E", hover_color="#1D8B4E", text_color="#FFFFFF", corner_radius=8)
            btn.grid(row=i//cols, column=i%cols, padx=2, pady=2, sticky="nsew")
            btn.is_cell = True
            box_buttons.append(btn)

    def update_stats(g_count, wins, last_opened, strat):
        nonlocal session_total_games, session_total_wins, session_total_losses
        import datetime
        
        fails = g_count - wins
        
        session_total_games += g_count
        session_total_wins += wins
        session_total_losses += fails
        
        success_rate = (session_total_wins / session_total_games) * 100 if session_total_games > 0 else 0
        
        lbl_total_games.configure(text=f"{T('total_games_res')}\n{session_total_games}")
        if success_rate.is_integer():
            lbl_success_rate.configure(text=f"{T('success_pct')}\n{int(success_rate)}%")
        else:
            lbl_success_rate.configure(text=f"{T('success_pct')}\n{success_rate:.1f}%")
        lbl_wins.configure(text=f"{T('success_res')}\n{session_total_wins}")
        lbl_losses.configure(text=f"{T('fail_res')}\n{session_total_losses}")
        
        for i, btn in enumerate(box_buttons):
            box_num = i + 1
            if box_num in last_opened:
                btn.configure(fg_color="#808080", hover_color="#808080")
            else:
                btn.configure(fg_color="#1D8B4E", hover_color="#1D8B4E")
                
        nonlocal is_simulating
        is_simulating = False
        btn_start.configure(state="normal")
        
        # Обновляем глобальную статистику
        StatsData.total_games += g_count
        if strat == "cycles":
            StatsData.cyclic_total += g_count
            StatsData.cyclic_success += wins
            StatsData.cyclic_fails += fails
        else:
            StatsData.random_total += g_count
            StatsData.random_success += wins
            StatsData.random_fails += fails
        StatsData.save_db()
            
        StatsData.success_games = StatsData.cyclic_success + StatsData.random_success
        if StatsData.total_games > 0:
            StatsData.success_rate = int((StatsData.success_games / StatsData.total_games) * 100)
            
        rate = int((wins / g_count) * 100) if g_count > 0 else 0
        now = datetime.datetime.now()
        strat_str = T("strategy_cyclic_hist") if strat == "cycles" else T("strategy_random_hist")
        StatsData.history.insert(0, {
            'is_win': wins > 0,
            'strategy': f"Авто: {strat_str}",
            'rate': rate,
            'date': now.strftime("%d.%m.%y"),
            'time': now.strftime("%H:%M:%S")
        })

    set_prisoners(100)
    set_games(100)
    set_strategy("cycles")

def back_to_game_screen_from_auto(frame):
    def _do():
        try:
            frame.destroy()
        except Exception:
            pass
        app.minsize(420, 600)
        show_game_screen()
    transition_to(_do)

# =============================================
# ЭКРАН СТАТИСТИКИ
# =============================================

def show_stats_screen(active_tab="stats"):
    _stop_particles()
    for w in app.winfo_children():
        if getattr(w, "is_overlay", False): continue
        try:
            w.place_forget()
            w.pack_forget()
        except Exception:
            pass

    app.minsize(900, 600)

    global stats_screen_frame
    if stats_screen_frame is not None:
        try:
            stats_screen_frame.destroy()
        except Exception:
            pass

    stats_screen_frame = ctk.CTkFrame(app, fg_color="#030303")
    stats_screen_frame.pack(fill="both", expand=True)

    top_bar = ctk.CTkFrame(stats_screen_frame, fg_color="transparent")
    top_bar.pack(fill="x", pady=10, padx=20)

    ctk.CTkLabel(top_bar, text=T("stats_title"),
                 font=ctk.CTkFont(family="Arial", size=28, weight="bold"),
                 text_color="#1eb852").pack(side="left")

    exit_btn_st = ctk.CTkButton(top_bar, text=T("btn_exit"),
                  font=ctk.CTkFont(family="Arial", size=20, weight="bold"),
                  text_color="#1eb852", fg_color="transparent",
                  hover_color="#1a1a1a", width=60,
                  command=lambda: back_to_menu(stats_screen_frame)
                  )
    exit_btn_st.pack(side="right")
    exit_btn_st.bind("<Enter>", lambda e: exit_btn_st.configure(text_color="#FF0000"))
    exit_btn_st.bind("<Leave>", lambda e: exit_btn_st.configure(text_color="#1eb852"))
    exit_btn_st.bind("<ButtonPress-1>", lambda e: exit_btn_st.configure(text_color="#FF0000"))

    ctk.CTkLabel(stats_screen_frame,
                 text=f"{T('stats_total')}: {StatsData.total_games}",
                 font=ctk.CTkFont(family="Arial", size=18, weight="bold"),
                 text_color="#1eb852", anchor="w").pack(fill="x", padx=20)

    tabs_frame = ctk.CTkFrame(stats_screen_frame, fg_color="transparent")
    tabs_frame.pack(fill="x", padx=20, pady=10)

    tab_stats = ctk.CTkLabel(tabs_frame, text=T("stats_tab_stats"),
                              font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                              text_color="#1eb852", cursor="hand2")
    tab_stats.pack(side="left", padx=(0, 20))

    tab_history = ctk.CTkLabel(tabs_frame, text=T("stats_tab_history"),
                                font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                                text_color="#888888", cursor="hand2")
    tab_history.pack(side="left")

    ctk.CTkFrame(stats_screen_frame, fg_color="#1eb852", height=2).pack(fill="x", padx=20, pady=(0, 15))

    content_frame = ctk.CTkFrame(stats_screen_frame, fg_color="transparent")
    content_frame.pack(fill="both", expand=True)

    stats_content   = ctk.CTkFrame(content_frame, fg_color="transparent")
    history_content = ctk.CTkScrollableFrame(content_frame, fg_color="transparent")

    def switch_tab(tab_name):
        if tab_name == "stats":
            tab_stats.configure(text_color="#1eb852")
            tab_history.configure(text_color="#888888")
            history_content.pack_forget()
            stats_content.pack(fill="both", expand=True)
        else:
            tab_stats.configure(text_color="#888888")
            tab_history.configure(text_color="#1eb852")
            stats_content.pack_forget()
            history_content.pack(fill="both", expand=True)

    tab_stats.bind("<Button-1>",   lambda e: switch_tab("stats"))
    tab_history.bind("<Button-1>", lambda e: switch_tab("history"))

    # --- STATS ---
    cards_frame = ctk.CTkFrame(stats_content, fg_color="transparent")
    cards_frame.pack(fill="x", padx=15, pady=5)
    cards_frame.grid_columnconfigure((0, 1, 2), weight=1)

    def create_card(parent, title, value, row, col):
        card = ctk.CTkFrame(parent, fg_color="#121212", corner_radius=10, height=100)
        card.grid(row=row, column=col, padx=10, pady=5, sticky="nsew")
        card.grid_propagate(False)
        ctk.CTkLabel(card, text=title,
                     font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                     text_color="#1eb852").place(x=15, y=15)
        ctk.CTkLabel(card, text=value,
                     font=ctk.CTkFont(family="Arial", size=36, weight="bold"),
                     text_color="#1eb852").place(x=15, y=45)

    create_card(cards_frame, T("stats_total"),   f"#{StatsData.total_games}",   0, 0)
    create_card(cards_frame, T("stats_success"), f"#{StatsData.success_games}", 0, 1)
    create_card(cards_frame, T("stats_rate"),    f"#{StatsData.success_rate}%", 0, 2)

    charts_frame = ctk.CTkFrame(stats_content, fg_color="transparent")
    charts_frame.pack(fill="x", padx=15, pady=10)
    charts_frame.grid_columnconfigure((0, 1), weight=1)

    chart1_frame = ctk.CTkFrame(charts_frame, fg_color="#121212", corner_radius=10, height=280)
    chart1_frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
    chart1_frame.pack_propagate(False)

    ctk.CTkLabel(chart1_frame, text=T("stats_compare"),
                 font=ctk.CTkFont(family="Arial", size=20, weight="bold"),
                 text_color="#1eb852").pack(anchor="w", padx=20, pady=15)

    chart1_canvas = ctk.CTkCanvas(chart1_frame, width=500, height=200, bg="#121212", highlightthickness=0)
    chart1_canvas.pack(padx=20, pady=(0, 15), fill="both", expand=True)

    max_val = max(100, StatsData.cyclic_total, StatsData.random_total)
    for i in range(5):
        y   = 20 + i * 40
        val = int(max_val - (max_val * i / 4))
        chart1_canvas.create_line(50, y, 480, y, fill="#333333", dash=(4, 4))
        chart1_canvas.create_text(45, y, text=str(val), fill="#1eb852", anchor="e", font=("Arial", 12))

    chart1_canvas.create_line(50, 180, 480, 180, fill="#1eb852")
    bar_w = 60
    for offset, succ, fail, label, lcolor in [
        (100, StatsData.cyclic_success, StatsData.cyclic_fails, T("stats_cyclic").replace("⟲ ", ""), "#1eb852"),
        (300, StatsData.random_success, StatsData.random_fails, T("stats_random"), "#e62e2e"),
    ]:
        h1 = 160 * (succ / max_val) if max_val > 0 else 0
        chart1_canvas.create_rectangle(offset, 180 - h1, offset + bar_w, 180, fill="#1eb852", outline="")
        h2 = 160 * (fail / max_val) if max_val > 0 else 0
        chart1_canvas.create_rectangle(offset + bar_w, 180 - h2, offset + bar_w * 2, 180, fill="#e62e2e", outline="")
        chart1_canvas.create_text(offset + bar_w, 195, text=label, fill=lcolor, anchor="n", font=("Arial", 12))

    chart2_frame = ctk.CTkFrame(charts_frame, fg_color="#121212", corner_radius=10, height=280)
    chart2_frame.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
    chart2_frame.pack_propagate(False)

    ctk.CTkLabel(chart2_frame, text=T("stats_overall"),
                 font=ctk.CTkFont(family="Arial", size=20, weight="bold"),
                 text_color="#1eb852").pack(anchor="w", padx=20, pady=15)

    chart2_canvas = ctk.CTkCanvas(chart2_frame, width=500, height=200, bg="#121212", highlightthickness=0)
    chart2_canvas.pack(padx=20, pady=(0, 15), fill="both", expand=True)

    cx, cy, r = 250, 100, 90
    chart2_canvas.create_arc(cx - r, cy - r, cx + r, cy + r, start=0, extent=360, fill="#e62e2e", outline="")
    if StatsData.total_games > 0:
        success_angle = int(360 * StatsData.success_games / StatsData.total_games)
        if success_angle > 0:
            chart2_canvas.create_arc(cx - r, cy - r, cx + r, cy + r,
                                      start=270, extent=success_angle,
                                      fill="#1eb852", outline="white", width=2)
    chart2_canvas.create_text(cx + 120, cy + 70,
                               text=f"{T('stats_success_label')} {StatsData.success_rate}%",
                               fill="#1eb852", font=("Arial", 16, "bold"))

    details_frame = ctk.CTkFrame(stats_content, fg_color="transparent")
    details_frame.pack(fill="x", padx=15, pady=5)
    details_frame.grid_columnconfigure((0, 1), weight=1)

    def add_detail_row(parent, label, value, clr):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(row, text=label, font=ctk.CTkFont(family="Arial", size=14), text_color=clr).pack(side="left")
        ctk.CTkLabel(row, text=value, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), text_color=clr).pack(side="right")

    dcard1 = ctk.CTkFrame(details_frame, fg_color="#121212", corner_radius=10)
    dcard1.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
    ctk.CTkLabel(dcard1, text=T("stats_cyclic"),
                 font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                 text_color="#1eb852").pack(anchor="w", padx=20, pady=15)
    add_detail_row(dcard1, T("stats_total_g"), f"#{StatsData.cyclic_total}",   "#1eb852")
    add_detail_row(dcard1, T("stats_succ"),    f"#{StatsData.cyclic_success}", "#1eb852")
    add_detail_row(dcard1, T("stats_fail"),    f"#{StatsData.cyclic_fails}",   "#1eb852")

    dcard2 = ctk.CTkFrame(details_frame, fg_color="#121212", corner_radius=10)
    dcard2.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
    ctk.CTkLabel(dcard2, text=T("stats_random"),
                 font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                 text_color="#e62e2e").pack(anchor="w", padx=20, pady=15)
    add_detail_row(dcard2, T("stats_total_g"), f"#{StatsData.random_total}",   "#e62e2e")
    add_detail_row(dcard2, T("stats_succ"),    f"#{StatsData.random_success}", "#e62e2e")
    add_detail_row(dcard2, T("stats_fail"),    f"#{StatsData.random_fails}",   "#e62e2e")

    # --- HISTORY ---
    for item in StatsData.history:
        clr        = "#1eb852" if item['is_win'] else "#e62e2e"
        title_text = T("history_win") if item['is_win'] else T("history_lose")

        card = ctk.CTkFrame(history_content, fg_color="transparent",
                             border_width=2, border_color=clr, corner_radius=15, height=100)
        card.pack(fill="x", padx=20, pady=10)
        card.pack_propagate(False)

        icon_f = ctk.CTkFrame(card, width=70, height=70, fg_color=clr, corner_radius=15)
        icon_f.pack(side="left", padx=15, pady=15)
        icon_f.pack_propagate(False)
        ctk.CTkLabel(icon_f, text="🎲", font=ctk.CTkFont(size=40)).place(relx=0.5, rely=0.5, anchor="center")

        res_f = ctk.CTkFrame(card, fg_color="transparent", width=120)
        res_f.pack(side="left", fill="y", pady=20)
        ctk.CTkLabel(res_f, text=title_text,
                     font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                     text_color=clr).pack(anchor="w")
        ctk.CTkButton(res_f, text=T("btn_again"),
                      font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                      text_color=clr, fg_color="transparent",
                      border_width=1, border_color=clr,
                      corner_radius=10, width=100, height=30).pack(anchor="w", pady=(5, 0))

        mid_f = ctk.CTkFrame(card, fg_color="transparent")
        mid_f.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        ctk.CTkLabel(mid_f,
                     text=f"{T('history_date')} {item['date']} | {T('history_time')} {item['time']}",
                     font=ctk.CTkFont(family="Arial", size=12),
                     text_color="#888888").pack(anchor="w")
        ctk.CTkLabel(mid_f, text=item['strategy'],
                     font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
                     text_color=clr).pack(anchor="w", pady=(5, 0))

        ctk.CTkLabel(card, text=f"{T('stats_success_label')} {item['rate']}%",
                     font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
                     text_color=clr).pack(side="right", padx=30)

    switch_tab(active_tab)


# =============================================
# ЗАПУСК
# =============================================

app.attributes("-alpha", 0.0)
build_menu()

if __name__ == "__main__":
    app.mainloop()